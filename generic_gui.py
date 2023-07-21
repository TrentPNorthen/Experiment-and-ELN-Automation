#Benchling GUI

import os
import tkinter as tk
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import requests
import pandas as pd
from itertools import compress, chain
import csv
import xlsxwriter
import datetime
from openpyxl import load_workbook
import string

############## Global Lists/Dictionaries #############
config = {}
cat_lip_names = []
help_lip_names = []
lipid_names_types = {}
lipid_types = []
mrna_exports = {}
lipid_exports = {}
process_exports = {}
doe_exports = []

############## Preset Dictionaries #######################

#Adds presets to lipids names types list for tab5
lipid_names_types['Lipid1'] = 'Cationic Lipid'
lipid_names_types['Lipid2'] = 'Cationic Lipid'
lipid_names_types['Lipid3'] = 'Cationic Lipid'
lipid_names_types['Phoslip1'] = 'Phoslipid'
lipid_names_types['DSPC'] = 'Phoslipid'

#presets for tab treeviews
lipid_presets = [
    ['Lipid2', '1000', 'LP-Lipid1_ID', 'Cationic Lipid'],
    ['Lipid1', '1100', 'LP-Lipid2_ID', 'Cationic Lipid'],
    ['Lipid3', '1200', 'LP-Lipid3_ID', 'Cationic Lipid']
    ]

helper_presets = [
    ]

mrna_presets = [
    ['EPO', 'EPO', '', '', ''],
    ['FFluc', 'FFluc', '', '', ''],
    ['mix1', 'mRNA1', '', '', ''],
    ['mix2', 'mRNA1', 'mRNA2', 'mRNA3', 'mRNA4']
    ]

process_presets = [
    ['Process1', '50', '300', '1', 'Formulate + Exchange1', 'Buffer 1'],
    ['Process2', '100', '350', '2', 'Formulate + Exchange2.', 'Buffer 2'],
    ['Process4.', '150', '400', '3', 'Formulate + Exchange 3', 'Buffer 3'],
    ['Process5', '200', '450', '4', 'Formulate + Exchange 4', 'Buffer 4'],
    ['Process3', '250', '500', '5', 'Formulate + Exchange 5', 'Buffer 5']
    ]

#mrna_mix presets for exports
mrna_mixes = {
    'Generic': {
        'mrna 1' : 'EPO',
        'mrna 2' : '',
        'mrna 3' : '',
        'mrna 4' : '',
        'mrna 5' : '',
        'mrna 6' : '',
        'mrna 7' : '',
        'mrna 8' : '',
        'mrna 9' : '',
        'mrna 10': '',
        'mrna 1 ratio': '1',
        'mrna 2 ratio': '',
        'mrna 3 ratio': '',
        'mrna 4 ratio': '',
        'mrna 5 ratio': '',
        'mrna 6 ratio': '',
        'mrna 7 ratio': '',
        'mrna 8 ratio': '',
        'mrna 9 ratio': '',
        'mrna 10 ratio': '',
        'mrna mixture id': 'MMI083',
        },
    
    'EPO' : {
        'mrna 1' : 'EPO',
        'mrna 2' : '',
        'mrna 3' : '',
        'mrna 4' : '',
        'mrna 5' : '',
        'mrna 6' : '',
        'mrna 7' : '',
        'mrna 8' : '',
        'mrna 9' : '',
        'mrna 10': '',
        'mrna 1 ratio': '1',
        'mrna 2 ratio': '',
        'mrna 3 ratio': '',
        'mrna 4 ratio': '',
        'mrna 5 ratio': '',
        'mrna 6 ratio': '',
        'mrna 7 ratio': '',
        'mrna 8 ratio': '',
        'mrna 9 ratio': '',
        'mrna 10 ratio': '',
        'mrna mixture id': 'MMI083',
              },
    
    'FFluc': {
        'mrna 1' : 'FFLuc',
        'mrna 2' : '',
        'mrna 3' : '',
        'mrna 4' : '',
        'mrna 5' : '',
        'mrna 6' : '',
        'mrna 7' : '',
        'mrna 8' : '',
        'mrna 9' : '',
        'mrna 10': '',
        'mrna 1 ratio': '1',
        'mrna 2 ratio': '',
        'mrna 3 ratio': '',
        'mrna 4 ratio': '',
        'mrna 5 ratio': '',
        'mrna 6 ratio': '',
        'mrna 7 ratio': '',
        'mrna 8 ratio': '',
        'mrna 9 ratio': '',
        'mrna 10 ratio': '',
        'mrna mixture id': 'insert ffluc id',
              },
    'mRNA1': {
        'mrna 1' : 'mRNA1',
        'mrna 2' : '',
        'mrna 3' : '',
        'mrna 4' : '',
        'mrna 5' : '',
        'mrna 6' : '',
        'mrna 7' : '',
        'mrna 8' : '',
        'mrna 9' : '',
        'mrna 10': '',
        'mrna 1 ratio': '1',
        'mrna 2 ratio': '',
        'mrna 3 ratio': '',
        'mrna 4 ratio': '',
        'mrna 5 ratio': '',
        'mrna 6 ratio': '',
        'mrna 7 ratio': '',
        'mrna 8 ratio': '',
        'mrna 9 ratio': '',
        'mrna 10 ratio': '',
        'mrna mixture id': 'MMI132'
              },
    'mRNA2': {
        'mrna 1' : 'mRNA1',
        'mrna 2' : 'mRNA2',
        'mrna 3' : 'mRNA3',
        'mrna 4' : 'mRNA4',
        'mrna 5' : '',
        'mrna 6' : '',
        'mrna 7' : '',
        'mrna 8' : '',
        'mrna 9' : '',
        'mrna 10': '',
        'mrna 1 ratio': '0.25',
        'mrna 2 ratio': '0.25',
        'mrna 3 ratio': '0.25',
        'mrna 4 ratio': '0.25',
        'mrna 5 ratio': '',
        'mrna 6 ratio': '',
        'mrna 7 ratio': '',
        'mrna 8 ratio': '',
        'mrna 9 ratio': '',
        'mrna 10 ratio': '',
        'mrna mixture id': 'MMI095'
              },
    }

#lipid mix presets for exports
lipid_mixes = {
    'Standard' : {
                  'Component 6': '',
                  'Component 7': '',
                  'Component 8': '',
                  'Component 9': '',
                  'Component 10': '',
                  'Lipid Mixture ID' : '',
                  'Cationic' : 10.0,
                  'Phospholipid' : 20.0,
                  'PEG' : 30.0,
                  'Cholesterol' : 40.0,
                  'Component 5 Ratio' : '',
                  'Component 6 Ratio' : '',
                  'Component 7 Ratio' : '',
                  'Component 8 Ratio' : '',
                  'Component 9 Ratio' : '',
                  'Component 10 Ratio' : '',
                  'Lipid Mixture ID': '',
                  },
    
    'Mix2' : {
                'Component 6': '',
                'Component 7': '',
                'Component 8': '',
                'Component 9': '',
                'Component 10': '',
                'Lipid Mixture ID' : '',
                'Cationic' : 20.0,
                'Phospholipid' : 30.0,
                'PEG' : 40.0,
                'Cholesterol' : 10.0,
                'Component 5 Ratio' : '',
                'Component 6 Ratio' : '',
                'Component 7 Ratio' : '',
                'Component 8 Ratio' : '',
                'Component 9 Ratio' : '',
                'Component 10 Ratio' : '',
                'Lipid Mixture ID' : '',
                },
    
    'Mix3' : {
                'Component 6': '',
                'Component 7': '',
                'Component 8': '',
                'Component 9': '',
                'Component 10': '',
                'Lipid Mixture ID' : '',
                'Cationic' : 40.0,
                'Phospholipid' : 10.0,
                'PEG' : 20.0,
                'Cholesterol' : 30.0,
                'Component 5 Ratio' : '',
                'Component 6 Ratio' : '',
                'Component 7 Ratio' : '',
                'Component 8 Ratio' : '',
                'Component 9 Ratio' : '',
                'Component 10 Ratio' : '',
                'Lipid Mixture ID': '',
                },

    'Comp. 5': {
                'Component 6': '',
                'Component 7': '',
                'Component 8': '',
                'Component 9': '',
                'Component 10': '',
                'Lipid Mixture ID' : '',
                'Cationic' : 30.0,
                'Phospholipid' : 40.0,
                'PEG' : 10.0,
                'Cholesterol' : 20.0,
                'Component 5 Ratio' : 5.0,
                'Component 6 Ratio' : '',
                'Component 7 Ratio' : '',
                'Component 8 Ratio' : '',
                'Component 9 Ratio' : '',
                'Component 10 Ratio' : '',
                'Lipid Mixture ID': '',
                },

    'doe'    : {
                'Component 6': '',
                'Component 7': '',
                'Component 8': '',
                'Component 9': '',
                'Component 10': '',
                'Lipid Mixture ID' : '',
                'Cationic' : 'doe cat',
                'Phospholipid' : 'doe phos',
                'PEG' : 'doe peg',
                'Cholesterol' : 'doe chol',
                'Component 5 Ratio' : '',
                'Component 6 Ratio' : '',
                'Component 7 Ratio' : '',
                'Component 8 Ratio' : '',
                'Component 9 Ratio' : '',
                'Component 10 Ratio' : '',
                'Lipid Mixture ID' : '',
                }
    }

#lipid parameters
lipid_params = {
    'Lipid2' : {
        'Molecular Weight' : 1000.0,
        'Benchling Lipid ID' : 'LP-Lipid2 Cationic -620003N20 -035',
        },
    'Lipid1' : {
        'Molecular Weight': 1100.0,
        'Benchling Lipid ID' : 'LP-Lipid1 -D-343-125 -1102',
        },
    'Lipid3' : {
        'Molecular Weight' : 1200.0,
        'Benchling Lipid ID' : 'LP-Lipid3-630002N20 -1103',
        },
    'Phospholip2' : {
        'Molecular Weight': 744.05,
        'Benchling Lipid ID' : 'LP-Phospholip2-CoE Formulations -1105',
        },
    'PEG 2k DMG' : {
        'Molecular Weight': 2468,
        'Benchling Lipid ID': 'LP-DMG-PEG-2000-CoE Formulations -1106',
        },
    'Cholesterol' : {
        'Molecular Weight': 386.5,
        'Benchling Lipid ID': 'LP-Cholesterol-CoE Formulations -1104',
        },
    'Phoslip1' : {
        'Molecular Weight': 856.25,
        'Benchling Lipid ID': 'LP-Phoslip1 Helper -CoE Formulations -1123',
        },
    'DSPC' : {
        'Molecular Weight': 790.15,
        'Benchling Lipid ID': 'LP-DSPC-CoE Formulations -1113'
        },
    'Lipid4': {
        'Molecular Weight': 1300.00,
        'Benchling Lipid ID': 'LP- IM-001_Lipid4 Cationic Isomannide-M1266-881193 -1855',
        }
    
        
    }

#Process presets for exports
processes = {
    'Generic' : {
                'Formulation Process': '',
                'N/P': '',
                'Organic Solvent': '',
                'Acidification Buffer': '',
                'Mixing Equipment': '',
                'Aqueous Flow Rate (mL/min)': '',
                'Lipid Flow Rate (mL/min)':'',
                'Buffer Exchange Equipment': '',
                'Buffer Exchange 1': '',
                'Buffer Exchange 2': '',
                'Buffer Exchange 3': '',
                'Final Buffer': '',
                'Process ID': ''
                       },
    
    'Process1' : {
                'Formulation Process': 'Conventional',
                'N/P': 4,
                'Organic Solvent': '100% Ethanol',
                'Acidification Buffer': 'Citrate 1mM NaCl 150mM pH4.5',
                'Mixing Equipment': 'HT Syringe',
                'Aqueous Flow Rate (mL/min)': 50,
                'Lipid Flow Rate (mL/min)': 300,
                'Buffer Exchange Equipment': 'AMBR',
                'Buffer Exchange 1': '20% EtOH',
                'Buffer Exchange 2': '10% Trehalose',
                'Buffer Exchange 3': '',
                'Final Buffer': '10% Trehalose',
                'Process ID': ''
                       },
    
    'Process2' : {
                'Formulation Process': 'Conventional',
                'N/P': 4,
                'Organic Solvent': '100% Ethanol',
                'Acidification Buffer': 'Citrate 1mM NaCl 150mM pH4.5',
                'Mixing Equipment': 'HT Syringe',
                'Aqueous Flow Rate (mL/min)': 100,
                'Lipid Flow Rate (mL/min)': 350,
                'Buffer Exchange Equipment': 'Dialysis/Amicon',
                'Buffer Exchange 1': '20% EtOH',
                'Buffer Exchange 2': 'Water',
                'Buffer Exchange 3': '10% Trehalose',
                'Final Buffer': '10% Trehalose',
                'Process ID': ''
                        },
    
    'Process4.' : {
                'Formulation Process': 'Conventional',
                'N/P': 4,
                'Organic Solvent': '100% Ethanol',
                'Acidification Buffer': 'Citrate 1mM NaCl 150mM pH4.5',
                'Mixing Equipment': 'Syringe',
                'Aqueous Flow Rate (mL/min)': 150,
                'Lipid Flow Rate (mL/min)': 400,
                'Buffer Exchange Equipment': 'Dialysis/Amicon',
                'Buffer Exchange 1': '20% EtOH',
                'Buffer Exchange 2': 'Water',
                'Buffer Exchange 3': '10% Trehalose',
                'Final Buffer': '10% Trehalose',
                'Process ID': '',
                        },
    
    'Process3' : {
                'Formulation Process': 'Conventional',
                'N/P': 4,
                'Organic Solvent': '100% Ethanol',
                'Acidification Buffer': 'Citrate 1mM NaCl 150mM pH4.5',
                'Mixing Equipment': 'Gear Pump',
                'Aqueous Flow Rate (mL/min)': 200,
                'Lipid Flow Rate (mL/min)': 450,
                'Buffer Exchange Equipment': 'TFF',
                'Buffer Exchange 1': '20% EtOH',
                'Buffer Exchange 2': '10% Trehalose',
                'Buffer Exchange 3': '',
                'Final Buffer': '10% Trehalose',
                'Process ID': '',
                        },

    'Process5' : {
                'Formulation Process': 'Conventional',
                'N/P' : 4,
                'Organic Solvent': '100% Ethanol',
                'Acidification Buffer': 'Citrate 1mM NaCl 150mM pH4.5',
                'Mixing Equipment': 'Syringe',
                'Aqueous Flow Rate (mL/min)': 250,
                'Lipid Flow Rate (mL/min)': 500,
                'Buffer Exchange Equipment': 'AMBR',
                'Buffer Exchange 1': '20% EtOH',
                'Buffer Exchange 2': '10% Trehalose',
                'Buffer Exchange 3': '',
                'Final Buffer': '10% Trehalose',
                'Process ID': '',
                        },
    }

#Used to create bencling ids for acidification buffers from shorthand
acid_buffers = {
    'Buffer1' : 'Component 1 + Component 2 + Component 3 pH1',
    'Buffer2' : 'Component 2 + Component 3 + Component 4 pH2'
    }

#Colum Labels for exports
columns_list = [
                'Formulation Date',
                'Benchling Formulation ID',
                'Formulation LOT #',
                'Batch Size (mg)',
                'mRNA 1',
                'mRNA 2',
                'mRNA 3',
                'mRNA 4',
                'mrna 5',
                'mrna 6',
                'mrna 7',
                'mrna 8',
                'mrna 9',
                'mrna 10',
                'mrna 1 ratio',
                'mrna 2 ratio',
                'mrna 3 ratio',
                'mrna 4 ratio',
                'mrna 5 ratio',
                'mrna 6 ratio',
                'mrna 7 ratio',
                'mrna 8 ratio',
                'mrna 9 ratio',
                'mrna 10 ratio',
                'mRNA Mixture ID',
                'Cat. Lip Name',
                'Cat. Lip #',
                'Lipid Lot #',
                'Ethanolic Solution Batch#',
                'Cat. Lip',
                'Phospholip',
                'PEG-like',
                'Chol-like',
                'Component 5',
                'Component 6',
                'Component 7',
                'Component 8',
                'Component 9',
                'Component 10',
                'Lipid Mixture ID',
                'Cationic (mol%)',
                'Phospolip (mol%)',
                'PEG-like (mol%)',
                'Chol-like (mol%)',
                'Component 5 (mol%)',
                'Component 6 (mol%)',
                'Component 7 (mol%)',
                'Component 8 (mol%)',
                'Component 9 (mol%)',
                'Component 10 (mol%)',
                'Formulation Process',
                'N/P',
                'Organic Solvent',
                'Acidification Buffer',
                'Mixing Equipment',
                'Aqueous Flow Rate (mL/min)',
                'Lipid Flow Rate (mL/min)',
                'Buffer Exchange Equiment',
                'Buffer Exchange 1',
                'Buffer Exchange 2',
                'Buffer Exchange 3',
                'Final Buffer',
                'Formulation Process ID',
                'Size (nm)',
                'PDI',
                'Encapsulation (%)',
                'mRNA Concentration (mg/mL)',
                'mRNA Integrity (%)',
                'Visual Appearance',
                'pH',
                'Zeta Potential (mV)',
                'Osmolality',
                'Component 1 amount (mg/mL)',
                'Component 2 amount (mg/mL)',
                'Component 3 amount (mg/mL)',
                'Component 4 amount (mg/mL)',
                'Component 5 amount (mg/mL)',
                'Component 6 amount (mg/mL)',
                'Component 7 amount (mg/mL)',
                'Component 8 amount (mg/mL)',
                'Component 9 amount (mg/mL)',
                'Component 10 amount (mg/mL)',
                                       ]

############## Business Logic #################### 

#adds lots to listbox
def append_lots():

    def concatenate_lots():
        for lot_no in lot_nos_list:
            lots_list.append(f'{_notebook_entry.get()}-{lot_no}')

    config['_formulation_lots'] = []
    lots_nos_list = [] #for numbers only
    lots_list = [] #notebook numbers + numbers
 
    start = int(_lot_nums_entry.get())
    amount = int(_count.get())
    final_lot_no = start + amount 

    lot_nos_list = list(range(start,final_lot_no)) #numbers only list
    concatenate_lots()

    if final_lot_no > 0:
            _lots_cumulative.set(lots_list)

    config['_formulation_lots'] = lots_list
    
#adds entries/dropdowns into treeview
def add_entries(self, _tab):

    #enters entries/dropdowns into tv
    def tab_tv_insert(self, _tab):
        self.tab_tv.insert(
            parent='', index='end', text='', values= (
            *[entry.get() for entry in config[f'{self}_entries']],
            *[dropdown.get() for dropdown in config[f'{self}_ddnames']]
        )
                       )

    #executes add lipids button for tab2
    def tab2_add(self, _tab):

        #Inserts entries/dropdowns into tab2 tv
        tab_tv_insert(self, _tab)

        #Adds names to dropdowns in _tab5
        try:
            if config[f'names_list{_tab}']:
                print('adding to list')
                pass

        except:
            print('names list does not exist')
            print('creating new list in tab2_add')
            config[f'names_list{_tab}'] = []
        
        config[f'names_list{_tab}'].append(config[f'{self}_entries'][0])
    ##############################    End of subfunctions     ##############################

    if _tab == 2:
        tab2_add(self, _tab)
        update_dds_list(self, _tab)

    elif _tab > 2 and _tab <= 4:
        tab_tv_insert(self, _tab)
        
        update_dds_list(self, _tab)
        update_dds(self, _tab)
        
    else:
        tab5_add(self)

    append_export_dict(self, _tab)

    #clears entry boxes after adding
    [entry.delete(0, END) for entry in config[f'{self}_entries']]

#Executes "Add Formulations" button for tab5
def tab5_add(self, func='add'):
        
    lipidmix = {}
    chosen_cat = config[f'{self}_ddnames'][1].get()
    chosen_help= config[f'{self}_ddnames'][2].get()

    lipidmix = {'Cationic':chosen_cat, 'Phoslipid':'Phospholip2', 'PEG':'PEG 2k DMG', 'Cholest': 'Cholesterol', 'Component 5': ''}

    if chosen_help:
        if lipid_names_types[chosen_help] == 'Phoslipid':
            lipidmix.update({'Phoslipid': chosen_help})
        if lipid_names_types[chosen_help] == 'PEG Lipid':
            lipidmix.update({'PEG': chosen_help})
        if lipid_names_types[chosen_help] == 'Cholesterol':
            lipidmix.update({'Cholest': chosen_help})
        if lipid_names_types[chosen_help] == 'Component 5':
            lipidmix.update({'Component 5': chosen_help})

    if func == 'add':
        lot_selector = config['_formulation_lots'][len(self.tab_tv.get_children())]
        
        self.tab_tv.insert(parent='', index='end', text='', values=(
            lot_selector,
            *[entry.get() for entry in config[f'{self}_entries']],
            config[f'{self}_ddnames'][0].get(),
            lipidmix['Cationic'],
            lipidmix['Phoslipid'],
            lipidmix['PEG'],
            lipidmix['Cholest'],
            lipidmix['Component 5'],
            *[dropdown.get() for dropdown in config[f'{self}_ddnames'][3:5]]
            )
                            )
    else:
        _selected = self.tab_tv.focus()
        lot_selector = config['_formulation_lots'][len(self.tab_tv.get_children())-1]

        self.tab_tv.item(_selected, text='', values=(
            lot_selector,
            *[entry.get() for entry in config[f'{self}_entries']],
            config[f'{self}_ddnames'][0].get(),
            lipidmix['Cationic'],
            lipidmix['Phoslipid'],
            lipidmix['PEG'],
            lipidmix['Cholest'],
            lipidmix['Component 5'],
            *[dropdown.get() for dropdown in config[f'{self}_ddnames'][3:5]]
            )
                           )

    #Retrieves dictionary values from items for export to spreadsheet
    lipid_exports[f'{lot_selector}'] = lipid_mixes[config['ddvars'][0].get()]
    mrna_exports[f'{lot_selector}'] = mrna_mixes[config['ddvars'][3].get()]
    process_exports[f'{lot_selector}'] = processes[config['ddvars'][4].get()]

    #Changes acidification buffer terminology for export to spreadsheet
    processes[config['ddvars'][4].get()]['Acidification Buffer'] = processes[
        config[f'{self}_ddnames'][4].get()]['Acidification Buffer']
    process_exports[f'{lot_selector}']['Acidification Buffer'] = processes[
        config['ddvars'][4].get()]['Acidification Buffer']
    

#adds lipids to respective lipid list; to be used for _tab5
def append_lipids(self):
    append_lip_types(self)
    
    if lipid_names_types[config[f'{self}_entries'][0].get()] == 'Cationic Lipid':
        cat_lip_names.append(config[f'{self}_entries'][0].get())
    else:
        help_lip_names.append(config[f'{self}_entries'][0].get())
        
def append_lip_types(self):
    lipid_names_types[config[f'{self}_entries'][0].get()] = config[f'{self}_ddnames'][0].get()
    lipid_types.append(config[f'{self}_ddnames'][0].get())

#appends a dictionary with data to be put into excel sheet
def append_export_dict(self, _tab):
    
    if _tab == 2:
        lipid_params[config[f'{self}_entries'][0].get()] = {
            'Lipid Name': config[f'{self}_entries'][0].get(),
            'Molecular Weight' : config[f'{self}_entries'][1].get(),
            'Benchling Lipid ID': config[f'{self}_entries'][2].get()
            }

    #same logic as _tab == 2 (above), except automated
    if _tab == 3:

        mrna_mixes[config[f'{self}_entries'][0].get()] = {}
        
        for n, item in enumerate(mrna_mixes['Generic']):
            if n <= 3:
                mrna_mixes[config[f'{self}_entries'][0].get()][item] = config[f'{self}_entries'][n+1].get()
            else:
                mrna_mixes[config[f'{self}_entries'][0].get()][item] = ''

        #Automatically adds same ratio for mrna ratios 1 through 4
        mrna_ratios = [
            'mrna 1 ratio',
            'mrna 2 ratio',
            'mrna 3 ratio',
            'mrna 4 ratio',
            ]

        mrna_count = []
        
        n = 1  
        for ratio in mrna_ratios:
            if not config[f'{self}_entries'][n].get() == '':
                mrna_count.append(f'mrna f{n} ratio')
            else:
                pass
            n += 1

        for ratio in mrna_ratios:
            mrna_mixes[config[f'{self}_entries'][0].get()][ratio] = 1/len(mrna_count)
            
    if _tab == 4:
        proc_name = config[f'{self}_entries'][0].get()
        processes[proc_name] = {}

        constants = {'Formulation Process': 'Conventional', 'Organic Solvent': '100% Ethanol', 'Process ID': ''}
        ents = ['Aqueous Flow Rate (mL/min)', 'Lipid Flow Rate (mL/min)', 'N/P']
        dds = ['Acidification Buffer']
        dicts = ['Mixing Equipment', 'Buffer Exchange Equipment', 'Buffer Exchange 1',
                 'Buffer Exchange 2', 'Buffer Exchange 3', 'Final Buffer']

        for parameter in constants:
            processes[proc_name][parameter] = constants[parameter]

        n=1
        for entry in ents:
            processes[proc_name][entry] = config[f'{self}_entries'][n].get()
            n+=1

        n=1
        for dd in dds:
            processes[proc_name]['Acidification Buffer']= config[f'{self}_ddnames'][n].get()
            n+=1

        for parameter in dicts:
            processes[proc_name][parameter] = processes[config[f'{self}_ddnames'][0].get()][parameter]

#grabs line and puts it into entry boxes
def select_entries(self, _tab):
    _selected = self.tab_tv.focus()

    #clear entry boxes and lists tv values
    [entry.delete(0, END) for entry in config[f'{self}_entries']]
    values = self.tab_tv.item(_selected,'values')
    val_list = list(values)

    #ensures lot is not appended to first entry box
    val_list1 = val_list[1:3]
    val_list2 = val_list[8:]
    lip_sorter = val_list[3:8]
    for lip in lip_sorter:
        if not lip == 'Phospholip2':
            if not lip == 'PEG 2k DMG':
                if not lip == 'Cholesterol':
                    if not lip == '':
                        print('c5', lip)
                        val_list1.append(lip)
        
                    else:
                        print('else lip', lip)
                        val_list1.append('')

    print('here', val_list1)
    val_list = val_list1 + val_list2

    #output to entry boxes
    n=0
    for entry in config[f'{self}_entries']:
        entry.insert(0, val_list[n])
        n += 1
        
    for dd in config[f'{self}_ddnames']:
        dd.set(val_list[n])
        n += 1

    #removes name from list that will go into tab5 dropdown
    if _tab == 2:
        if val_list[3] == 'Cationic Lipid':
            print('removing cationic lipid')
            cat_lip_names.remove(values[0])
        else:
            print('removing helper lipid')
            help_lip_names.remove(val_list[0])

        del lipid_params[config[f'{self}_entries'][0].get()]
            
    #if tab 5 does not have a names list
    else:
        if _tab == 5:
            pass
        else:
            if _tab == 3:
                del mrna_mixes[values[0]]
            if _tab == 4:
                del processes[values[0]]
                
                
            config[f'names_list{_tab}'].remove(values[0])

    update_dds(self, _tab)

#inputs new values
def update_entry(self, _tab):
    _selected = self.tab_tv.focus()

    #save new data
    if not _tab == 5:
        self.tab_tv.item(_selected, text="", values=(
        *[entry.get() for entry in config[f'{self}_entries']],
        *[dropdown.get() for dropdown in config[f'{self}_ddnames']])
                           )
    else:
        tab5_add(self, func = 'update')

    #adds new entry to list which will go into tab5
    if _tab == 2:
        append_lipids(self)

        #updates lipid parameters
        lipid_params[config[f'{self}_entries'][0].get()] = {
        'Molecular Weight' : config[f'{self}_entries'][1].get(),
        'Benchling Lipid ID' : config[f'{self}_entries'][2].get(),
        }
        
    elif _tab == 5:
        pass
    
    else:
        config[f'names_list{_tab}'].append(config[f'{self}_entries'][0].get())
        append_export_dict(self, _tab)


    #clear entry boxes
    [entry.delete(0, END) for entry in config[f'{self}_entries']]

    update_dds_list(self, _tab, func = 'update')
    update_dds(self, _tab)
    
#clears values in the treeview and relavent lists
def clear_entries(self, _tab):
    self.tab_tv.delete(*self.tab_tv.get_children())

    if _tab == 2:
        del (cat_lip_names[:], help_lip_names[:])
        
    else:
        try:
            del config[f'names_list{_tab}'][:]

        except:
            print('List is empty')
            pass
    
    update_dds_list(self, _tab)      
    update_dds(self, _tab)

#takes imported values from table, and puts them in a list
def import_names_list(self, _tab):
    rows = []
    for row_id in self.tab_tv.get_children():
        row = list(self.tab_tv.item(row_id, 'values'))
        rows.append(row)
        
        config[f'names_list{_tab}'].append(row[0])
    
#updates dropdown contents in tab5 with added/deleted
def update_dds_list(self, _tab, func = 'add'):
  
    _selected = self.tab_tv.focus()

    if _tab == 2:
        if func == 'add':
            append_lipids(self)
            update_dds(self, _tab)
        else:
            append_lip_types(self)

    else:
        print('not tab2')
        
        try:
            #note: names list below is always deleted when importing
            if config[f'names_list{_tab}']:
                print('adding to list')
                pass

        except:
            print('names list does not exist')
            print('creating new list in update dds')
            config[f'names_list{_tab}'] = []
            
        finally:
            #this portion is here for updating list when adding with the add button
            if not config[f'{self}_entries'][0].get() == '':
                config[f'names_list{_tab}'].append(config[f'{self}_entries'][0].get())

            #this portion is here for the updating list when importing
            else:
                if func == 'add':
                    import_names_list(self, _tab)
                else:
                    pass
                
#updates tab5 dropdowns if mrna or process was imported
def import_dds_list(self, _tab):
    import_names_list(self, _tab)
    update_dds(self, _tab)    

#updates actual dropdown items
def update_dds(self, _tab):

    def tab5_dds_appends(self, _tab):
        #m is used so presets aren't deleted from tab 5 when adding
        for dd in config['dds']:
            self.input_dropdown = config['dds'][n]["menu"]
            config['dds'][n]["menu"].delete(m ,"end")
        
        for name in config[f'names_list{_tab}']:
            self.input_dropdown.add_command(label=name,
                command=lambda value=name: config['ddvars'][n].set(value))
            
    if _tab == 2:
        if lipid_names_types[config[f'{self}_entries'][0].get()] == 'Cationic Lipid':
            lip_type = 'cationic'
            input_cat_dd = config['dds'][1]["menu"]
            delete_lip_dds(self, 'input_cat_dd', input_cat_dd, lip_type)
        else:
            lip_type = 'helper'
            input_help_dd = config['dds'][2]["menu"]
            delete_lip_dds(self, 'input_help_dd', input_help_dd, lip_type)
        append_lip_dds(lip_type)
 
    if _tab == 3:
        n = 3
        m = len(mrna_presets)
        tab5_dds_appends(self, _tab)

    if _tab == 4:
        n = 4
        m = len(process_presets)
        tab5_dds_appends(self, _tab)

#updates dropdowns when importing tab2
def import_tab2_dds(self, _tab):
    
    #this is for when lipids are imported
    for lipids in cat_lip_names:
        lip_type = 'cationic'
        input_cat_dd = config['dds'][1]["menu"]
        delete_lip_dds(self, 'input_cat_dd', input_cat_dd, lip_type)
    append_lip_dds(lip_type)
        
    for lipids in help_lip_names:
        lip_type = 'helper'
        input_help_dd = config['dds'][2]["menu"]
        delete_lip_dds(self, 'input_help_dd', input_help_dd, lip_type)
    help_lip_names.insert(0, '')
    append_lip_dds(lip_type)
    
#deletes previous option menus to remove duplicates from appending
def delete_lip_dds(self, name, dd, lip_type):
    if lip_type == 'cationic':
        m = len(lipid_presets)
    if lip_type == 'helper':
        m = len(helper_presets)+1

    try:
        config[name].delete(m, "end")
    except:
        print('Dropdown does not exist')

    #configures outside of this function 
    try:
        config[name] = dd
    except:
        pass

#adds new dropdown list to tab5
def append_lip_dds(lip_type):

    if lip_type == 'cationic':
        for cat_lip in cat_lip_names[1:]:
            config['input_cat_dd'].add_command(label=cat_lip,
                            command=lambda value=cat_lip: config['ddvars'][1].set(value))
    else:
        for help_lip in help_lip_names[0:]:
            try:
                config['input_help_dd'].add_command(label=help_lip,
                        command=lambda value=help_lip: config['ddvars'][2].set(value))
            except:
                print('Helper lipid list not created; only cationic lipids have been input')

#allows for import of excel file into tab treeviews
def import_tv(self, _type, _tab):
    
    filename = filedialog.askopenfilename(
        initialdir="/Users/trentnorthen/Desktop/LPP1/Sanofi/benchling/benchling_gui",
        title = "Select A File",
        filetypes=(("xlsx files", "*.xlsx"), ("All Files", "*.*"))
        )

    if filename:
        try:
            filename = r"{}".format(filename)
            read_file = pd.read_excel(filename)
        except ValueError:
            my_label.config(text="File Could Not Be Opened")
        except FileNotFoundError:
            my_label.config(text="File Could Not be Found")

        if not _type == 'doe':
            clear_entries(self, _tab)

    if _type == 'tab':

        #setup new treeview from import
        read_rows = read_file.to_numpy().tolist()
        for row in read_rows:
            self.tab_tv.insert("", "end", values=row)

        if _tab == 2:
            for row in self.tab_tv.get_children():
                values = list(self.tab_tv.item(row, 'values'))
                lipid_names_types[values[0]] = values[3]

                lipid_params[values[0]] = {}
                
                lipid_params[values[0]]['Molecular Weight'] = values[1]
                lipid_params[values[0]]['Benchling Lipid ID'] = values[2]
                
                #Adds lipid to respective dropdown in popup
                if lipid_names_types[values[0]] == 'Cationic Lipid':      
                    cat_lip_names.append(values[0])
                else:
                    help_lip_names.append(values[0])

            update_dds(self, _tab)
            import_tab2_dds(self, _tab)

        if _tab == 3:
            for row in self.tab_tv.get_children():
                values = list(self.tab_tv.item(row, 'values'))

                #Used to automatically create equal ratios of mrnas
                ratio = []
                r = 1
            
                for value in values[1:]:
                    if not value == 'nan':
                        ratio.append(r)
                        r += 1
                    else:
                        pass

                mrna_mixes[values[0]] = {
                    'mrna 1' : values[1],
                    'mrna 2' : values[2],
                    'mrna 3' : values[3],
                    'mrna 4' : values[4],
                    'mrna 5' : '',
                    'mrna 6' : '',
                    'mrna 7' : '',
                    'mrna 8' : '',
                    'mrna 9' : '',
                    'mrna 10': '',
                    'mrna 1 ratio': '',
                    'mrna 2 ratio': '',
                    'mrna 3 ratio': '',
                    'mrna 4 ratio': '',
                    'mrna 5 ratio': '',
                    'mrna 6 ratio': '',
                    'mrna 7 ratio': '',
                    'mrna 8 ratio': '',
                    'mrna 9 ratio': '',
                    'mrna 10 ratio': '',
                    'mrna mixture id': '',
                    }

                #automatically inputs ratios, assumes all are equal
                for n in ratio:
                    mrna_mixes[values[0]]['mrna ' f'{n}' ' ratio'] = 1/len(ratio)

            import_dds_list(self, _tab)

        if _tab == 4:
            for row in self.tab_tv.get_children():
                values = list(self.tab_tv.item(row, 'values'))

                processes[values[0]] = {}
                constants = {'Formulation Process':'Conventional', 'Organic Solvent':'100% Ethanol', 'Process ID': ''}
                vals = ['Aqueous Flow Rate (mL/min)', 'Lipid Flow Rate (mL/min)', 'N/P']
                dicts = ['Mixing Equipment', 'Buffer Exchange Equipment', 'Buffer Exchange 1', 'Buffer Exchange 2',
                         'Buffer Exchange 3', 'Final Buffer', 'Acidification Buffer']

                for parameter in constants:
                    processes[values[0]][parameter] = constants[parameter]

                n=1
                for parameter in vals:
                    processes[values[0]][parameter] = values[n]
                    n+=1

                for parameter in dicts:
                    processes[values[0]][parameter] = processes[values[4]][parameter]

            import_dds_list(self, _tab)

    if _type == 'popup':
        #setup new treeview from import
        read_rows = read_file.to_numpy().tolist()
        for row in read_rows:
            self.popup_tv.insert("", "end", values=row)

    if _type == 'doe':
        read_rows = read_file.to_numpy().tolist()
        for row in read_rows:
            self.doe_tv.insert("", "end", values=row)
      
#appends information from tabs into popup tv
def append_popup(self, _tab, names_list=[]):
    
    lot_count = []

    #The difference between this and cat_lip_names list is this accounts for count differences by adding ''
    cat_lip_appends = []
    help_lip_appends = []
    tv_appends = []

    n = 0
    m = 0

    for formulation_lot in config['_formulation_lots']: 
        lot_count.append(formulation_lot)

    #inserts lots from lots tab into popup treeview
    if _tab == 2:

        while len(cat_lip_appends+help_lip_appends) < len(config['_formulation_lots']):

            #if no helper lipid entered, assumes defaults
            for cat_lip in cat_lip_names[1:]:
                cat_lip_appends.append(cat_lip)
                help_lip_appends.append('')

            #if screening helper lipids, it assumes Lipid1
            #extra space is inside when importing for dd, we don't want to append
            if help_lip_names:
                if help_lip_names[0] == '':
                    for help_lip in help_lip_names[1:]:
                        cat_lip_appends.append('Lipid1')
                        help_lip_appends.append(help_lip)
                else:
                    for help_lip in help_lip_names:
                        cat_lip_appends.append('Lipid1')
                        help_lip_appends.append(help_lip)

            if len(cat_lip_appends+help_lip_appends) >= len(config['_formulation_lots']):
                print('breaking')
                break


    #adds placeholder into help lip dropdown if not testing new helper lipids
        if len(help_lip_names) == 0:
            print('help lip length is zero')
            help_lip_names.append('')
        else:
            pass
    
    #accounts for count differences
        while len(cat_lip_appends+help_lip_appends) >= len(config['_formulation_lots']):
            for formulation_lot in config['_formulation_lots'][m:len(config['_formulation_lots'])+1]:
                cat_lip_appends.append('')
                help_lip_appends.append('')
                m += 1
            else:
                break

        n = 0
        while len(self.popup_tv.get_children()) < len(config['_formulation_lots']):
            for formulation_lot in config['_formulation_lots']:
                self.popup_tv.insert( 
                    parent='', index='end',iid=n,text='',
                    values=(formulation_lot, 1, 'Standard',
                            cat_lip_appends[n], help_lip_appends[n], 'EPO', 'Process1'))
                n += 1
                
            else:
                break

    else: 
        while len(tv_appends) < len(config['_formulation_lots']):
            for mrna in config[f'names_list{_tab}']:
                tv_appends.append(mrna)

    #accounts for count differences
        while len(tv_appends) >= len(config['_formulation_lots']):
            for formulation_lot in config['_formulation_lots'][m:len(config['_formulation_lots'])+1]:
                tv_appends.append('')
                m += 1
            else:
                break

        n = 0
        if _tab == 3:
            for formulation_lot in config['_formulation_lots']:
                self.popup_tv.insert( 
                    parent='', index='end',iid=n,text='',
                    values=(formulation_lot, 1, 'Standard', 'Lipid1', '', tv_appends[n], 'Process3'))
                n += 1

        if _tab == 4:
            for formulation_lot in config['_formulation_lots']:
                self.popup_tv.insert( 
                    parent='', index='end',iid=n,text='',
                    values=(formulation_lot, 1, 'Standard', 'Lipid1', '', 'EPO', tv_appends[n]))
                n += 1

#Popup Functions, these are separate from tab functions, except for export which also applies to tab5

def apply(self, items = []):

    try:
        selected = self.popup_tv.focus()
    except:
        selected = self.doe_tv.focus()
   
    try:
        values = list(self.popup_tv.item(selected, 'values'))
    except:
        values = list(self.doe_tv.item(selected, 'values'))
    
    #blanks in dropdown menus will not change the corresponding values in treeview
    for n in list(range(6)):
        if items[n].get() == '':
            pass
        else:
            values[n+1] = items[n].get()
    
    apply = [values[0], *values[1:]]

    if selected:
        self.popup_tv.item(selected, text="", values= apply)

def apply_all(self, items = []):
            
    for row in self.popup_tv.get_children():
        values = list(self.popup_tv.item(row, 'values'))

        for n in list(range(6)):
            if items[n].get() == '':
                pass
            else:
                values[n+1] = items[n].get()
        
        apply = [values[0], *values[1:]]
        self.popup_tv.item(row, text="", values= apply)
        
#exports to spreadsheet
def export(self, _type, cols_list = []):

    form_path = r"/Users/trentnorthen/Desktop/LPP1/Sanofi/benchling/benchling_gui/Formulation_Template.xlsx"
    excel_name = f"{config['_formulation_lots'][0]} through {config['_formulation_lots'][-1]}.xlsx"  

    rows = []
    
    book = load_workbook(form_path)
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')

    #creates a csv file, which can be exported to excel
    #csv file is separate from excel, and is named "Benchling_csv"
    ben_path = "/Users/username/Desktop/test_csv"
    temp_path = "/Users/username/Desktop/template_csv"
    lip_path = "/Users/username/Desktop/lipid_csv"

    paths = [
        ben_path,
        temp_path,
        lip_path,
        ]

    #deletes old csv file
    for path in paths:
        if os.path.exists(path):
            print('deleting')
            os.remove(path)
        else:
            print('files dont exist')

    #activates benchling worksheet
    ws = book['Benchling']
    
    with open('Benchling_csv', "w", newline = '') as myfile:
        csvwriter = csv.writer(myfile, delimiter=',')

        if _type == _tab:
            tab_export(self, _type, csvwriter=csvwriter,  rows=rows, cols_list=cols_list)

        elif _type == doe:
            doe_export(self, _type, csvwriter=csvwriter,  rows=rows, cols_list=cols_list)

        else:
            screen_export(self, _type, csvwriter=csvwriter,  rows=rows, cols_list=cols_list)

    df = pd.read_csv('Benchling_csv')

    #i is the csv row number, and row is the dict item + key
    for i, row in df.iterrows():

        n=0
        for item in cols_list:
            ws.cell(row=1, column=n+1).value = cols_list[n]
            n+=1
        
        for c, value in enumerate(row, start=0): #enumerates dict keys
            ws.cell(row=i+2, column=c+1).value = value
        i+=1
       
    df.to_excel(writer, 'Benchling')

    ws = book['Lipid Parameters']
    rows = []
    cols_list = ['Benchling Lipid ID', 'Lipid Name', 'Molecular Weight (g/mol)']
    rows.append(cols_list)

    with open("Lipid_csv", "w", newline='') as myfile:
        csvwriter = csv.writer(myfile, delimiter=',')
        for lipid in lipid_params:
            rows.append(
                [
                    lipid_params[f'{lipid}']['Benchling Lipid ID'],
                    lipid,
                    int(lipid_params[f'{lipid}']['Molecular Weight'])
                    ]
                )
        for row in rows: 
            csvwriter.writerow(row)
            
    df = pd.read_csv('Lipid_csv')
    
    #i is the csv row number, and row is the dict item + key
    for i, row in df.iterrows():
        
        for c, value in enumerate(cols_list, start=1): #enumerates list
            ws.cell(row=i+1, column=c).value = value
        i += 1
        
        for c, value in enumerate(row, start=0): #enumerates dict keys
            ws.cell(row=i, column=c+1).value = value
        i+=1

    df.to_excel(writer, 'Lipid Parameters')
    writer.book = book
    print('Creating Spreadsheet!')
    writer.close()

def tab_export(self, _type, csvwriter, rows, cols_list):

    print('exporting tab')

    #total items, items to be deleted (replaced with dict values)
    tot_itm = len(self.tab_tv['columns'])
    itm_tbd = 8

    for row_id in self.tab_tv.get_children():
        row = list(self.tab_tv.item(row_id, 'values')) #list of items in each row of tree            
        rows.append(row)
        
    exports_list = [
        mrna_exports,
        lipid_exports,
        process_exports
        ]

    #fun little recursive function :)
    #creates a list of all exports
    for exports in exports_list:
        n=0
        for row in exports:
            if exports == mrna_exports:
                m = tot_itm - itm_tbd 
            else:
                m = len(rows[n]) - itm_tbd
                
            for label in exports[row]:
                rows[n].insert(m, exports[row][label])
                m += 1
            n += 1

    date = datetime.datetime.now()

    #columns in spreadsheet which will be empty
    blanks = [1, 26, 27, 28]
    
    #adds dictionary items/blanks into list to be exported to spreadsheet
    #could automate, but this is more readable
    for row in rows:
        row.insert(0, date.strftime("%x"))
        row.insert(24, row[-(itm_tbd-1)])
        row.insert(25, lipid_params[row[-(itm_tbd-1)]]['Benchling Lipid ID'])
        row.insert(26, lipid_params[row[-(itm_tbd-2)]]['Benchling Lipid ID'])
        row.insert(27, lipid_params[row[-(itm_tbd-3)]]['Benchling Lipid ID'])
        row.insert(28, lipid_params[row[-(itm_tbd-4)]]['Benchling Lipid ID'])
        row.insert(29, lipid_params[row[-(itm_tbd-5)]]['Benchling Lipid ID'])

        for blank in blanks:
            row.insert(blank, '')
    
    #deletes dictionary keys from list
    tbd = list(range(itm_tbd+1))[1:]
    for row in rows:
        for n in reversed(tbd):
            del row[-n]

    #writes csv file
    rows.insert(0, cols_list)
    for row in rows:
        csvwriter.writerow(row)

def doe_export(self, _type, csvwriter, rows, cols_list):

    print('exporting doe!')

    #total items, items to be deleted
    #Number of columns in treeview that will be replaced with dict items
    tot_itm = len(self.doe_tv['columns'])    
    itm_tbd = 8
    
    for row_id in self.doe_tv.get_children():
        row = list(self.doe_tv.item(row_id, 'values'))

        entries = [entry.get() for entry in config[f'{self}_entries']]
        dropdown = [dropdown.get() for dropdown in config[f'{self}_dropdowns']]

        row =  [row[0]] + entries + row[1:] + dropdown
        rows.append(row)
                
    n=0
    for row in rows:
        mrna_exports[row[0]] = mrna_mixes[row[8]]
        m = len(row) - itm_tbd
        for label in mrna_exports[row[0]]:
            rows[n].insert(m, mrna_exports[row[0]][label])
            m += 1
        n += 1

    n=0          
    for row in rows:

        #sets up lipid export dict
        lipid_exports[row[0]] = {
            'Cationic' : '',
            'Cat. Lip. #': '',
            'Lipid LOT #': '',
            'Ethanolic Solution Batch': '',
            'Cat. Lip ID': '', 
            'Phoslipid' : '',
            'PEG' : '',
            'Cholest' : ''
            }

        #defaults lipid settings
        lipid_exports[row[0]]['Cationic'] = row[-4]
        lipid_exports[row[0]]['Cat. Lip ID'] = lipid_params[row[-4]]['Benchling Lipid ID']
        lipid_exports[row[0]]['Phoslipid'] = 'LP-Phospholip2-CoE Formulations -1105'
        lipid_exports[row[0]]['PEG'] = 'LP-DMG-PEG-2000-CoE Formulations -1106'
        lipid_exports[row[0]]['Cholest'] = 'LP-Cholesterol-CoE Formulations -1104'
        lipid_exports[row[0]]['Component 5'] = ''

        #updates lipids based on selected helper lipid
        if not row[-3] == '':  
            if lipid_names_types[row[-3]]:
                if lipid_names_types[row[-3]] == 'Phoslipid':
                    lipid_exports[row[0]]['Phoslipid'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'PEG Lipid':
                    lipid_exports[row[0]]['PEG'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'Cholesterol':
                    lipid_exports[row[0]]['Cholest'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'Component 5':
                    lipid_exports[row[0]]['Component 5'] = lipid_params[row[-3]]['Benchling Lipid ID']
        else:
            print('Using default helper lipids')

    for row in rows:
        l = len(row) - itm_tbd

        for label in lipid_exports[row[0]]:
            rows[n].insert(l, lipid_exports[row[0]][label])
            l+=1

        for label in lipid_mixes['doe']:
            lipid_mixes['doe']['Cationic'] = row[len(row) - itm_tbd] 
            lipid_mixes['doe']['Phospholipid'] = row[len(row) - itm_tbd + 1]
            lipid_mixes['doe']['PEG'] = row[len(row) - itm_tbd + 2]
            lipid_mixes['doe']['Cholesterol'] = row[len(row) - itm_tbd + 3]
            lipid_mixes['doe']['Component 5'] = row[len(row) - itm_tbd + 4]
            rows[n].insert(l, lipid_mixes['doe'][label])
            l+=1
        n += 1
        
    n=0
    for row in rows:
        process_exports[row[0]] = processes[row[-1]]
        p = len(row) - itm_tbd
        for label in process_exports[row[0]]:
            rows[n].insert(p, process_exports[row[0]][label]) 
            p += 1
        n += 1

    tbd = list(range(itm_tbd+1))[1:]
    for row in rows:
        for n in reversed(tbd):
            del row[-n]     

    date = datetime.datetime.now()
    for row in rows:
        row.insert(0, date.strftime("%x"))
        row.insert(1, '')
                
    rows.insert(0, cols_list)
    for row in rows:
        csvwriter.writerow(row)

#exports screens popups
def screen_export(self, _type, csvwriter, rows, cols_list):
    print('exporting screen')

    #total columns, columns to be deleted, remaining columns
    #Number of columns in treeview that will be replaced with dict items
    tot_itm = len(self.popup_tv['columns'])    
    itm_tbd = 4

    for row_id in self.popup_tv.get_children():
        row = list(self.popup_tv.item(row_id, 'values')) #list of items in each row of tree
        rows.append(row)
            
    #Substitutes dropdown selection with extended values                
    #Change to function if you want; not as easy as it seems
                
    n=0
    for row in rows:
        mrna_exports[row[0]] = mrna_mixes[row[5]]
        m = tot_itm - itm_tbd - 1
        for label in mrna_exports[row[0]]:
            rows[n].insert(m, mrna_exports[row[0]][label])
            m += 1
        n += 1

    n=0
    for row in rows:
        lipid_exports[row[0]] = {
            'Cationic' : '',
            'Cat. Lip. #': '',
            'Lipid LOT #': '',
            'Ethanolic Solution Batch': '',
            'Cat. Lip ID': '', 
            'Phoslipid' : '',
            'PEG' : '',
            'Cholest' : '',
            'Component 5': '',
            }

        #defaults lipid settings
        lipid_exports[row[0]]['Cationic'] = row[-4]
        lipid_exports[row[0]]['Cat. Lip ID'] = lipid_params[row[-4]]['Benchling Lipid ID']
        lipid_exports[row[0]]['Phoslipid'] = 'LP-Phospholip2-CoE Formulations -1105'
        lipid_exports[row[0]]['PEG'] = 'LP-DMG-PEG-2000-CoE Formulations -1106'
        lipid_exports[row[0]]['Cholest'] = 'LP-Cholesterol-CoE Formulations -1104'
        lipid_exports[row[0]]['Component 5'] = ''

        #updates lipids based on selected helper lipid
        if not row[-3] == '':  
            if lipid_names_types[row[-3]]:
                if lipid_names_types[row[-3]] == 'Phoslipid':
                    lipid_exports[row[0]]['Phoslipid'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'PEG Lipid':
                    lipid_exports[row[0]]['PEG'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'Cholesterol':
                    lipid_exports[row[0]]['Cholest'] = lipid_params[row[-3]]['Benchling Lipid ID']
                if lipid_names_types[row[-3]] == 'Component 5':
                    lipid_exports[row[0]]['Component 5'] = lipid_params[row[-3]]['Benchling Lipid ID']
        else:
            print('Using default helper lipids')

    for row in rows:
        l = len(row) - itm_tbd - 1

        for label in lipid_exports[row[0]]:
            rows[n].insert(l, lipid_exports[row[0]][label])
            l+=1

        for label in lipid_mixes[row[-5]]:
            rows[n].insert(l, lipid_mixes[row[-5]][label])
            l+=1

        n += 1

    n=0
    for row in rows:
        process_exports[row[0]] = processes[row[-1]]
        p = len(row) - itm_tbd - 1
        for label in processes['Generic']:
            rows[n].insert(p, process_exports[row[0]][label]) 
            p += 1
        n += 1
    
    tbd = list(range(itm_tbd+2))[1:]
    for row in rows:
        for n in reversed(tbd):
            del row[-n]  

    date = datetime.datetime.now()
    for row in rows:
        row.insert(0, date.strftime("%x"))
        row.insert(1, '')
                
    rows.insert(0, cols_list)
    for row in rows:
        csvwriter.writerow(row)
        
  
############# Popups ######################
def popup(self, title, _tab, variable_list):

    _screen = tk.Toplevel()
    _screen.wm_title(title)

    #Popup Treeview
    self.popup_tv = ttk.Treeview(_screen, columns=(1,2,3,4,5), height=3) 
    self.popup_tv.grid(row=0, column=0, sticky=(E, W), columnspan=7)

    _scrollbar = ttk.Scrollbar(
        _screen, orient=VERTICAL, command=self.tab_tv.yview)
    _scrollbar.grid(row=0, column=7, sticky=(S, N), pady=6)

    columns = ['LOT', 'Batch Size', 'Lipid Mix', 'Cationic Lipid', 'Helper Lipid', 'mRNA', 'Process']
    texts = ["LOT", "Batch Size (mg)", "Lipid Mix", "Cat. Lipid", "Hel. Lipid", "mRNA", "Process"]
    cols_txts = zip(columns, texts)

    self.popup_tv['columns'] = ([column for column in columns])
    self.popup_tv.column("#0", width=0, stretch=NO)
    self.popup_tv.heading("#0", text="", anchor=CENTER)

    for column, text in cols_txts:
        self.popup_tv.column(column, anchor=CENTER, width=100)
        self.popup_tv.heading(column, text=text, anchor=CENTER)

    #Automatically insert tab info into popup Treeview
    itemized_tree_list = []
    names_list = []

    # m = _tab + 2 by coincidence, so will not automate
    n = 0
    if _tab == 2:
        m = 4
    if _tab == 3:
        m = 5
    if _tab == 4:
        m = 6
        
    for line in self.tab_tv.get_children(): 
        for value in self.tab_tv.item(line)['values'][n::m]:
            itemized_tree_list.append(value)
    
    while n < len(itemized_tree_list):
        name = itemized_tree_list[n]
        names_list.append(name)
        n += 1

    config['names_t'f'{_tab}'] = names_list
    append_popup(self, _tab, names_list = config['names_t'f'{_tab}'])

    #Popup Labels
    lbls = ["Batch Size (mg)", "Lipid Mix", "Cationic Lipid", "Helper Lipid", "mRNA", "Process"]
    n = list(range(6))
    lbls_cols = zip(lbls, n)

    for lbl, col in lbls_cols:
        label=ttk.Label(_screen, text=lbl, padding='2 2 0 0')
        label.grid(row=2, column=col, sticky=(E,W))

    #Popup Buttons/Dropdowns
    stringvars = []

    _export = ttk.Button(_screen, text='Create!', width=5, command=lambda: export(self, popup, columns_list))
    _export.grid(row=1, column=6, sticky=(E))

    _size_entry = ttk.Entry(_screen, width=5)
    _size_entry.grid(row=3, column=0, sticky=(E, W))
    stringvars.append(_size_entry)

    mx_prsts = ["", "Standard", "Mix2", "Mix3", "Comp. 5"]
 
    _lipid_mix = StringVar(_screen)
    _lipid_mix.set('Standard')
    stringvars.append(_lipid_mix)

    _lipid = StringVar(_screen)
    _lipid.set('')
    stringvars.append(_lipid)

    _help_lip = StringVar(_screen)
    _help_lip.set('')
    stringvars.append(_help_lip)

    _select_mrna = StringVar(_screen)
    _select_mrna.set('')
    stringvars.append(_select_mrna)

    _select_process = StringVar(_screen)
    _select_process.set('Process1')
    stringvars.append(_select_process)

    cat_lip_presets = ['', 'Lipid1','Lipid2','Lipid3','Lipid4']
    help_lip_presets = ['Phoslip1','DSPC']
    mrna_presets = ['', 'EPO','FFluc','mRNA1','mRNA2']
    process_presets = ['', 'Process3', 'Process1', 'Process5',' Process4.']

    #these lipid are inserted into the process dropdowns for all screens
    if cat_lip_names:
        cat_inputs = cat_lip_names + cat_lip_presets
    else:
        cat_inputs = cat_lip_presets

    if help_lip_names:
        help_inputs = help_lip_names + help_lip_presets
    else:
        help_inputs = help_lip_presets

    dd_strngs = [_lipid_mix, _lipid, _help_lip]
    dd_inpts = [mx_prsts, cat_inputs, help_inputs]
    strng_inpts = zip(dd_strngs, dd_inpts)

    n = 1
    for strng, inpt in strng_inpts:
        dropdown = OptionMenu(_screen, strng, *inpt)
        dropdown.grid(row=3, column=n, sticky=(E,W))
        n += 1

    #these mrna are inserted into the process dropdown if the tab is 3
    if _tab == 3:
        _input_mrna_dropdown = OptionMenu(_screen, _select_mrna, *config['names_t3'])
        _input_mrna_dropdown.grid(row=3, column=4, sticky=(E,W))
        
    #these mrna are inserted into the process dropdown if the tab is not 3
    else:
        try:
            del inputs[:]
            inputs = config['names_list3'] + mrna_presets
        except:
            print('names_list3 does not exist')
            inputs = mrna_presets
        
        _input_mrna_dropdown = OptionMenu(_screen, _select_mrna,  *inputs)
        _input_mrna_dropdown.grid(row=3, column=4, sticky=(E,W))

    #these processes are inserted into the process dropdown if the tab is 4
    if _tab == 4:
        _input_process_dropdown = OptionMenu(_screen, _select_process, *config['names_t4'])
        _input_process_dropdown.grid(row=3, column=5, sticky=(W))

    #these processes are inserted into the process dropdown if tab is not 4
    else:
        try:
            inputs = config['names_list4'] + process_presets
        except:
            print('names_list4 does not exist')
            inputs = process_presets
            
        _input_process_dropdown = OptionMenu(_screen, _select_process, *inputs)
        _input_process_dropdown.grid(row=3, column=5, sticky=(E,W))

    _apply = ttk.Button(_screen, text='Apply', command=lambda: apply(self, items = stringvars))
    _apply.grid(row=3, column=6, sticky=(E,W))

    _apply_all = ttk.Button(_screen, text='Apply to all', command=lambda: apply_all(self, items = stringvars))
    _apply_all.grid(row=4, column=6, sticky=(E,W))

    for variable in variable_list:
        variable.set(False)

#Creation of DOE tab, separate layout from screening tabs
def doe(self, title, _tab, variable_list):
    print('its doe time')
    _screen = tk.Toplevel()
    _screen.wm_title(title)

    config[f'{self}_entries'] = []
    entries = []
    dropdowns = []

    #Popup Treeview
    self.doe_tv = ttk.Treeview(_screen, columns=(1,2,3,4,5), height=3) 
    self.doe_tv.grid(row=0, column=0, sticky=(E, W), columnspan=6)

    _scrollbar = ttk.Scrollbar(
        _screen, orient=VERTICAL, command=self.tab_tv.yview)
    _scrollbar.grid(row=0, column=6, sticky=(S, N), pady=6)
 
    columns = ['LOT', 'PEG', 'Cat.', 'Chol.', 'Phospholip2', 'Comp. 5']
    texts = ["LOT", "PEG (%)", "Cat (%)", "Chol (%)", "Phosph (%)"]
    cols_txts = zip(columns, texts)

    self.doe_tv['columns'] = ([column for column in columns])
    self.doe_tv.column("#0", width=0, stretch=NO)
    self.doe_tv.heading("#0", text="", anchor=CENTER)

    for column, text in cols_txts:
        self.doe_tv.column(column, anchor=CENTER, width=100)
        self.doe_tv.heading(column, text=text, anchor=CENTER)
        
    #Popup Labels
    lbls = ["Batch Size (mg)", "Select Cat. Lip.", "Select Help. Lip.", "Select mRNA", "Select Process"]

    n=0
    for lbl in lbls:
        label = ttk.Label(_screen, text=lbl, padding='2 2 0 0')
        label.grid(row=2, column=n, sticky=(E,W))
        n+=1

    #Popup Buttons/Dropdowns
    stringvars = []

    _import = ttk.Button(_screen, text='Import', width=5, command=lambda: import_tv(self, 'doe', _tab))
    _import.grid(row=1, column=0, sticky=(W))

    _export = ttk.Button(_screen, text='Create!', width=5, command=lambda: export(self, doe, columns_list))
    _export.grid(row=1, column=5, sticky=(E))

    _size_entry = ttk.Entry(_screen, width=5)
    _size_entry.grid(row=3, column=0, sticky=(E, W))
    stringvars.append(_size_entry)
    entries.append(_size_entry)

    _lipid = StringVar(_screen)
    _lipid.set('')
    stringvars.append(_lipid)
    dropdowns.append(_lipid)

    _help_lip = StringVar(_screen)
    _help_lip.set('')
    stringvars.append(_help_lip)
    dropdowns.append(_help_lip)

    names_list = []
    itemized_tree_list = []

    for lipid in names_list:
        if lipid_names_types[lipid] == 'Cationic Lipid':
            cat_lip_names.append(lipid)
        else:
            pass

    config['names_t2'] = names_list

    cat_lip_presets = [
                'Lipid1',
                'Lipid2',
                'Lipid3',
                'Lipid4',
                ]

    cat_inputs = cat_lip_presets + cat_lip_names

    _input_lipid_dropdown = OptionMenu(_screen, _lipid, *cat_inputs)
    _input_lipid_dropdown.grid(row=3, column=1, sticky=(E,W))


    if not len(help_lip_names) == 0:
        _input_helper_dropdown = OptionMenu(_screen, _help_lip, '', *help_lip_names)
        _input_helper_dropdown.grid(row=3, column=2, sticky=(E,W))

    else:
        _input_helper_dropdown = OptionMenu(_screen, _help_lip, '')
        _input_helper_dropdown.grid(row=3, column=2, sticky=(E,W))

    #Uses default mrna and processes
    _select_mrna = StringVar(_screen)
    _select_mrna.set('')
    dropdowns.append(_select_mrna)

    _select_process = StringVar(_screen)
    _select_process.set('Process1')
    dropdowns.append(_select_process)
    
    _input_mrna_dropdown = OptionMenu(_screen, _select_mrna, 'EPO', 'FFluc', 'mRNA2', 'mRNA1')
    _input_mrna_dropdown.grid(row=3, column=3, sticky=(E,W))

    _input_process_dropdown = OptionMenu(_screen, _select_process, 'Process1', 'Process5', 'Process3')
    _input_process_dropdown.grid(row=3, column=4, sticky=(E,W))

    #adds buttons
    for variable in variable_list:
        variable.set(False)

    #for exports
    config[f'{self}_entries'] = entries
    config[f'{self}_dropdowns'] = dropdowns

############### Tab Layout Logic #######################
class Tab_Contents(Frame):

    #frame creation
    def __init__(self, _tab):
        Frame.__init__(self, _tab) #how does this work?

        self._tv_frame = ttk.Frame(_tab, padding='5 5 5 5')
        self._tv_frame.grid(row=0, column=0, sticky=(E, W))

        self._entry_frame = ttk.Frame(_tab, padding = '2 2 5 0')
        self._entry_frame.grid(row=2, column=0, sticky=(E, W))

        self._button_frame = ttk.Frame(_tab, padding = '2 2 5 0')
        self._button_frame.grid(row=3, column=0, sticky=(E,W))

    #creates treeview table
    def create_tv(self, _tab, col_names = [], tv_presets = []):

        if not _tab == 5:
            self.tab_tv = ttk.Treeview(self._tv_frame, columns=col_names, height=4)
            self.tab_tv.grid(row=0, column=0, columnspan = len(col_names), sticky=(E, W))

        else:
            self.tab_tv = ttk.Treeview(self._tv_frame, columns=col_names, height=5)
            self.tab_tv.grid(row=0, column=0, columnspan = len(col_names), sticky=(E, W))    

        self.tab_tv['columns'] = col_names

        self.tab_tv.column("#0", width=0, stretch=NO)
        self.tab_tv.heading("#0", text="", anchor=CENTER)

        self.tv_scrollbar = ttk.Scrollbar(
            self._tv_frame, orient=VERTICAL, command=self.tab_tv.yview)
        self.tv_scrollbar.grid(row=0, column=len(col_names), sticky=(S, N), pady=6)
        
        for each_col in col_names:
            self.tab_tv.column(
                each_col, anchor=CENTER, width=int(165 * 4 /len(col_names)))
            self.tab_tv.heading(each_col, text=each_col, anchor=CENTER)

        if not _tab == 5:    
            _upload_button = ttk.Button(self._tv_frame, text='Import', width=5, command=lambda: [import_tv(self, 'tab', _tab)])
            _upload_button.grid(row=1, column=(len(col_names)-1), sticky=(E))
            
        #inserts commonly used entities into treeview
        n=0
        for line in tv_presets:
            self.tab_tv.insert(parent='', index='end',iid=n,text='',values=line)
            n+=1

    #creates labels
    def create_labels(self, _tab, lab_names = []):
        
        n = 0
        for each_lab in lab_names:

            if not _tab == 5:
                _label_name = ttk.Label(self._entry_frame, text=each_lab, padding = '2 2 5 0')
                _label_name.grid(row=0, column=n, sticky=(E, W))
                n+=1

            if _tab == 5:
                _label_name = ttk.Label(self._entry_frame, text=each_lab, padding = '2 10 4 0')
                _label_name.grid(row=0, column=n, sticky=(E, W))
                n+=1
            
    #creates entry and dropdown boxes
    def create_entries(self, inputs, entry_cols=[]):

        config[f'{self}_entries'] = []
        entries = []

        n = 0
        for entry_col in entry_cols:
            _entry = ttk.Entry(self._entry_frame, width=int(65/inputs))
            _entry.grid(row=1, column=entry_col, sticky=(E, W))

            entries.append(_entry)
            n += 1

        config[f'{self}_entries'] = entries

    def create_dropdowns(self, dd_cols=[], dd_sets=[], dd_values=[]):
    
        dd_names = []
        dds = []
        ddvars = []

        n = 0
        for each_col in dd_cols:
            _ddvar = StringVar(self._entry_frame)
            _ddvar.set(dd_sets[n])
            ddvars.append(_ddvar)

            _dropdown = ttk.OptionMenu(
            self._entry_frame, _ddvar, *dd_values[n]
            ) 
            _dropdown.grid(row=1, column=dd_cols[n], sticky=(E,W))

            dd_names.append(_ddvar)
            dds.append(_dropdown)
            n+=1

        config[f'{self}_ddnames'] = dd_names
        config[f'dds'] = dds
        config['ddvars'] = ddvars

    def create_buttons(self, _tab, button_text =[], checkbox_text = [], button_commands = [], popup_list = []):

        n = 0
        for button in button_text:
            self._input_button = ttk.Button(
                self._button_frame, text=button, width=13, command=(lambda n=n : button_commands[n](self, _tab))
                )
            self._input_button.grid(row=2, column=n, sticky=(E,W))
            n += 1

        #allows checkboxes to be checked independently
        variable1 = tk.BooleanVar(value=False) 
        variable2 = tk.BooleanVar(value=False)

        variable_list = []
        variable_list.append(variable1)
        variable_list.append(variable2)

        m=3
        if not _tab == 5:
            for each_checkbox in checkbox_text:
                check = Checkbutton(
                    self._button_frame,
                    variable=variable_list[m-3],
                    text=each_checkbox,
                    padx = 10,
                    command= lambda m=m: popup_list[m-3](self, 'So you want to screen', _tab, variable_list))
                check.grid(row=m, column=3, sticky = (W))
                m += 1

            #_advanced = ttk.Button(
                #self._button_frame, text='Advanced', width=8, command = lambda: advanced())
            #_advanced.grid(row=4, column=0, sticky=(W), pady=5, padx=5)

        else:
            _export = ttk.Button(
                self._button_frame, text='Create!', width=13, command=(
                    lambda: export(self, _tab, cols_list = columns_list
                                       )
                    )
                )
            _export.grid(row=3, column=3, sticky=(E,W))
            
############### Creation of Tabs ################
if __name__ == "__main__":
    _root = Tk()
    _root.title('FormulaTron')
    _root.geometry("770x350")
    
#Tab contents
    _tab_frame = ttk.Frame(_root, padding='5 5 5 5')
    _tab_frame.grid(row=0, column=0, sticky=(S, N, E, W))
    _tabcontrol = ttk.Notebook(_tab_frame)
    _tab1 = ttk.Frame(_tabcontrol)
    _tab2 = ttk.Frame(_tabcontrol)
    _tab3 = ttk.Frame(_tabcontrol)
    _tab4 = ttk.Frame(_tabcontrol)
    _tab5 = ttk.Frame(_tabcontrol)
    _tab6 = ttk.Frame(_tabcontrol)
    _tabcontrol.add(_tab1, text='Notebook')
    _tabcontrol.add(_tab2, text='Lipid Entries')
    _tabcontrol.add(_tab3, text='mRNA Entries')
    _tabcontrol.add(_tab4, text='Process Entries')
    _tabcontrol.add(_tab5, text='Formulations')

    _tabcontrol.grid(row=0, column=0, sticky=(S, N, E, W))


#Tab 1
    _entry_frame_t1 = ttk.Frame(_tab1, padding='5 5 5 5')
    _entry_frame_t1.grid(row=0, column=0, sticky=(E, W, N, S))

    #Left Frame 
    _notebook_frame = ttk.LabelFrame(
        _entry_frame_t1, text="Notebook Number")
    _notebook_frame.grid(row=0, column=0, sticky=(E, W))
    
    _notebook = StringVar()
    _notebook.set('D680') #Change this for each lab member or find a way to automate
    _notebook_entry = ttk.Entry(
        _notebook_frame, width=10, textvariable=_notebook)
    _notebook_entry.grid(row=0, column=0, sticky=(E,W))

    _lot_nums_frame = ttk.LabelFrame(
        _entry_frame_t1, text="Starting LOT Number")
    _lot_nums_frame.grid(row=1, column=0, sticky=(E, W))
        
    _lot_nums = IntVar() #Restrict to numbers
    _lot_nums_entry = ttk.Entry(
        _lot_nums_frame, width=10, textvariable=_lot_nums,)
    _lot_nums_entry.grid(row=0, column=0, sticky=(E,W))
    _lot_nums.set('')

    _count_frame = ttk.LabelFrame(
        _entry_frame_t1, text="Number of Formulations")
    _count_frame.grid(row=2, column=0, sticky=(E, W))
        
    _count = IntVar() #Restrict to numbers
    _count_entry = ttk.Entry(
        _count_frame, width=10, textvariable=_count)
    _count_entry.grid(row=0, column=0, sticky=(E,W))
    _count.set('')

    _generate_lots_btn = ttk.Button(
        _entry_frame_t1, text='Generate LOTs', command=append_lots)
    _generate_lots_btn.grid(row=3, column=0, sticky=(E,W), pady=8)

    #Right Frame
    _lots_label_frame = ttk.LabelFrame(
        _entry_frame_t1, text='LOTS', padding='9 9 9 9')
    _lots_label_frame.grid(row=0, column=1, sticky=(N, S, E, W), rowspan=3) 

    _lots_cumulative = IntVar()
    
    _lots_cumul_listbox = Listbox(
        _lots_label_frame, listvariable=_lots_cumulative, height=8, width=30)
    _lots_cumul_listbox.grid(column=2, rowspan=3)
    _lots_cumulative.set('')

    _scrollbar = ttk.Scrollbar(
        _lots_label_frame, orient=VERTICAL, command=_lots_cumul_listbox.yview)
    _scrollbar.grid(row=0, column=4, sticky=(S, N))
    _lots_cumul_listbox.configure(yscrollcommand=_scrollbar.set)

#Lipids Tab
    _Tab2 = Tab_Contents(_tab2)
    _tab = 2

    _Tab2.create_tv(
        _tab,
        col_names = [
            'Lipid Name',
            'Molecular Weight',
            'Benchling Lipid ID',
            'Lipid Type' 
            ],
        tv_presets = lipid_presets
        )

    _Tab2.create_labels(
        _tab,
        lab_names = [
            'Lipid Name',
            'Molecular Weight',
            'Benchling Lipid ID',
            'Lipid Type'
            ]
        )

    _Tab2.create_entries(
        4,
        entry_cols = list(range(3)),
        )

    _Tab2.create_dropdowns(
        dd_cols = [
            3
            ],
        dd_sets = [
            "Cationic"
            ],
        dd_values = [
            [
            "Cationic Lipid",
            "Cationic Lipid",
            "Phoslipid",
            "PEG Lipid",
            "Cholesterol",
            "Component 5",
            ]
            ]
        )

    _Tab2.create_buttons(
        _tab,
        button_text = [
            'Add Lipid',
            'Edit Lipid',
            'Update Lipid',
            'Clear Table',
            ],
        checkbox_text = [
            'Lipid Screen?',
            'DOE'
            ],
        button_commands = [
            add_entries,
            select_entries,
            update_entry,
            clear_entries],
        popup_list = [
            popup,
            doe
            ]
        )
        
#mRNA Tab
    _Tab3 = Tab_Contents(_tab3)
    _tab = 3

    _Tab3.create_tv(
        _tab,
        col_names = [
            'Mix Name',
            'mRNA 1',
            'mRNA 2',
            'mRNA 3',
            'mRNA 4',
            ],
        tv_presets = mrna_presets
        )

    _Tab3.create_labels(
        _tab,
        lab_names = [
            'Mix Name',
            'mRNA 1',
            'mRNA 2',
            'mRNA 3',
            'mRNA 4',
            ]
        )

    _Tab3.create_entries(
        5,
        entry_cols = list(range(5)),
        )

    _Tab3.create_dropdowns(
        dd_cols = [
            ],
        dd_sets = [ 
            ],
        dd_values = [
            ]
        )

    _Tab3.create_buttons(
        _tab,
        button_text = [
            'Add mRNA',
            'Edit mRNA',
            'Update mRNA',
            'Clear Table',
            ],
        checkbox_text = [
            'mRNA Screen?',
            ],
        button_commands = [
            add_entries,
            select_entries,
            update_entry,
            clear_entries
            ],
        popup_list = [
            popup,
            ]
        )

#Formulations Tab

    _Tab4 = Tab_Contents(_tab4)
    _tab = 4

    _Tab4.create_tv(
        _tab,
        col_names = [
            'Process Name',
            'Aqueous Flow',
            'Organic Flow',
            'N/P',
            'T-Mix + Exchange',
            'Acidic Buffer',
            ],
        tv_presets = process_presets

        )

    _Tab4.create_labels(
        _tab,
        lab_names = [
            'Proc. Name',
            'Aq. Flow',
            'Org. Flow',
            'N/P',
            'T-Mix + Exch.',
            'Acidic Buffer',
            ]
        )

    _Tab4.create_entries(
        6,
        entry_cols = list(range(4)),
        )

    _Tab4.create_dropdowns(
        dd_cols = [
            4,
            5,
            ],
        dd_sets = [
            'HT Syr/AMBR',
            'Buffer2 shortened name',
            ],
        dd_values = [
            [
            'Process1',
            'Process2',
            'Process3',
            'Process4.',
            'Process5',
            ],
        [
            'Buffer2',
            'Buffer1',
            'Buffer2',
            ]
                     ]
        )

    _Tab4.create_buttons(
        _tab,
        button_text = [
            'Add Process',
            'Edit Process',
            'Update Process',
            'Clear Process',
            ],
        checkbox_text = [
            'Process Screen?',
            ],
        button_commands = [
            add_entries,
            select_entries,
            update_entry,
            clear_entries
            ],
        popup_list = [
            popup,
            ]
        )

#Tab 5
    _Tab5 = Tab_Contents(_tab5)
    _tab = 5

    _Tab5.create_tv(
        _tab,
        col_names = [
            'LOT',
            'Size (mg)',
            'Lipid Mix',
            'Cat. Lip.',
            'Phospholip.',
            'PEG-like',
            'Chol-like',
            'Comp. 5',
            'mRNA Mix',
            'Process'
            ],
        tv_presets = []
        )

    _Tab5.create_labels(
        _tab,
        lab_names = [
            'Size (mg)',
            'Lipid Mix',
            'Cat. Lip.',
            'Helper Lip.',
            'mRNA Mix',
            'Process'
            ]
        )

    _Tab5.create_entries(
        6,
        entry_cols = [0],
        )

    _Tab5.create_dropdowns(
        dd_cols = [
            1,
            2,
            3,
            4,
            5,
            ],
        
        dd_sets = [
            'Standard',
            "",
            "",
            'EPO',
            'Process1'
            ],
        
        dd_values = [

        [
            'Standard',
            'Standard', 
            'Mix2',
            'Mix3',
            'Comp. 5',
            ],
                     
        [
            'Lipid1',
            'Lipid1',
            'Lipid2',
            'Lipid3',
            ],

        [
            "",
            ""
            ],

        [
            'mRNA1',
            'EPO',
            'mRNA1',
            'mRNA2',
            'FFluc',
            ],

        [
            'Process1',
            'Process1',
            'Process2',
            'Process3',
            'Process4.',
            'Process5',
            ]
                     ]
        )
    
    _Tab5.create_buttons(
        _tab,
        button_text = [
            'Add Formulation',
            'Edit Formulation',
            'Update Formulation',
            'Clear Formulation',
            ],
        checkbox_text = [
            'Create!',
            ],
        button_commands = [
            add_entries,
            select_entries,
            update_entry,
            clear_entries]
        )

    _root.mainloop()

#To Do:

    ###########EASY##########
    #lipid import/addition isnt working after doe (small problem)
    #when creating spreadsheet, I get warning about book. Update.
    
    #when updating, the order changes for dds and screens. input into the same spot.
    #text stays when selecting blank on popup dropdowns
    #extra space in help dropdown after importing (cant change from [1:])
    #need to make new spreadsheet template with lip 5

   #########DIFFICULT#######
    #allow user to choose what folder to save exported data
    #User can save different things and then they stay in presets/dropdown if common

    #"Add new" in respective dropdown menus, which allows user to input their own parameters
    #let them create own lipid mixture % (Advanced?)
    #Add in for other lipid components (Advanced?)
    #Add in other process parameters (Advanced?)
    #Add in other mrna/ratios (Advanced?)

   ##########FINAL###########
    #Status bar with instructions?
    #Add in exception notifications
    #when edit is clicked, there should be a way to ensure they click update next (or, if they do something else, exception)

    
        
    
    
